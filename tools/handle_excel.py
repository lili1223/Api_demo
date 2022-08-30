"""
!/usr/bin/python3
 -*- coding: utf-8 -*-
@Time : 2022/8/19 20:49
@Project:py_Api
@Author : RainBow
@File : handle_excel.py
"""
from tools.handle_path import *
import os
from pprint import pprint
from openpyxl import load_workbook
class ReadExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)
        self.sheet_obj = self.wb[sheet_name]
    def get_data(self):
        """
        第一步: 获取指定表单的所有数据,存放在列表
        第二步：将测试数据和表头字段一一对应
        第三步：返回测试数据
        :return:
        """
        data = list(self.sheet_obj.iter_rows(values_only=True))
        title = data[0]
        cases = data[1:]
        cases_list = []
        for case in cases:
            res = dict(zip(title, case))
            cases_list.append(res)
        self.__close_excel()
        return cases_list
    def get(self):
        # 获取指定单元格的数据
        print(self.sheet_obj['A1'].value)
        print(self.sheet_obj.cell(1,2).value)
    # def write_excel(self,value):
    #     self.sheet_obj['A5'] = value
    #     self.sheet_obj.cell(4,6,'pyhhh')
    #     self.wb.save(self.file_name)
    def __close_excel(self):
        self.wb.close()

if __name__ == '__main__':
    res = ReadExcel(file_name=case_data_dir, sheet_name='login').get_data()
    print(res)
    # pprint(res)
    # res.get()
    # res.write_excel('hello')
